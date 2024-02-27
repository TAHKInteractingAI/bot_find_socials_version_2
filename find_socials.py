import tkinter as tk
import requests
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import filedialog
import re
import json
import time
import g4f
from parsel import Selector

g4f.debug.logging = True  # Enable logging
g4f.check_version = False  # Disable automatic version checking
import re
import spacy
# from spacy.cli.download import download
from g4f import Provider

attributes = dir(Provider)

methods = [
    attribute for attribute in attributes if callable(getattr(Provider, attribute))
]

# Filter out only methods starting with an uppercase letter (assumption: your methods start with an uppercase letter)
method_names = [method for method in methods if method[0].isupper()]

# Create a list of method references
method_references = [getattr(Provider, method) for method in method_names]

SOCIAL_MEDIA_PLATFORMS = ["Twitter", "Facebook", "LinkedIn"]

def clean_input(input_str):
    # Loại bỏ các ký tự đặc biệt, chỉ giữ lại ký tự chữ cái và dấu cách
    cleaned_str = re.sub(r"[^a-zA-Z\s]", "", input_str)
    return cleaned_str

def select_file():
    file_path = filedialog.askopenfilename(filetypes=(("All files", "*"),))
    file_entry.delete(0, tk.END)
    file_entry.insert(0, file_path)

def login_linkedin(driver):
    driver.get("https://www.linkedin.com/login")
    time.sleep(5)

    # Enter your LinkedIn login credentials
    username = driver.find_element(By.ID, "username")
    username.send_keys("thanhnhan.qn2002@gmail.com")  # Thay bằng địa chỉ email của bạn

    password = driver.find_element(By.ID, "password")
    password.send_keys("ThanhNhan2802.")  # Thay bằng mật khẩu của bạn

    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    time.sleep(5)

def extract_information_linkedin(driver, profile_links):
    login_linkedin(driver)
    linkedin_data = []

    for linkedin_url in profile_links:
        driver.get(linkedin_url)
        time.sleep(5)

        sel = Selector(text=driver.page_source)

        name_linkedin = sel.xpath('//*[starts-with(@class, "text-heading-xlarge")]/text()').extract_first()
        job_title_linkedin = sel.xpath('//*[starts-with(@class, "text-body-medium")]/text()').extract_first()
        followers_linkedin = sel.xpath('//*[starts-with(@class, "t-bold")]/text()').extract_first()
        location_linkedin = sel.xpath('//*[starts-with(@class, "text-body-small")]/text()').extract_first()

        linkedin_data.append({
            "Name LinkedIn": name_linkedin.strip() if name_linkedin else None,
            "Job Title": job_title_linkedin.strip() if job_title_linkedin else None,
            "Followers": followers_linkedin.strip() if followers_linkedin else None,
            "Location": location_linkedin.strip() if location_linkedin else None,
            "LinkedIn URL": driver.current_url,
        })

    return linkedin_data

def extract_infomation_twitter(driver, profile_links):
    twitter_data = []

    for twitter_url in profile_links:

        # get the profile URL 
        driver.get(twitter_url)

        # add a 5 second pause loading each URL
        time.sleep(5)

        # assigning the source code for the webpage to variable sel
        sel = Selector(text=driver.page_source)

        try:
            profile_name = sel.css('div.css-175oi2r.r-1awozwy.r-18u37iz.r-dnmrzs span.css-1qaijid.r-bcqeeo.r-qvutc0.r-poiln3:nth-child(1)::text').extract_first()
            profile_handle = sel.css('div.css-1rynq56.r-dnmrzs.r-1udh08x.r-3s2u2q.r-bcqeeo.r-qvutc0.r-37j5jr.r-a023e6.r-rjixqe.r-16dba41.r-18u37iz.r-1wvb978 span.css-1qaijid.r-bcqeeo.r-qvutc0.r-poiln3::text').extract_first()
            profile_job_title = sel.css('div.css-175oi2r.r-1adg3ll.r-6gpygo div.css-175oi2r div.css-1rynq56.r-bcqeeo.r-qvutc0.r-37j5jr.r-a023e6.r-rjixqe.r-16dba41 span.css-1qaijid.r-bcqeeo.r-qvutc0.r-poiln3::text').extract_first()
            profile_joining_date = sel.css('div.css-1rynq56.r-bcqeeo.r-qvutc0.r-37j5jr.r-a023e6.r-16dba41.r-56xrmm span.css-1qaijid.r-bcqeeo.r-qvutc0.r-poiln3.r-4qtqp9.r-1b7u577 span.css-1qaijid.r-bcqeeo.r-qvutc0.r-poiln3::text').extract_first()
            profile_following = sel.css('div.css-175oi2r.r-ymttw5.r-ttdzmv.r-1ifxtd0 div.css-175oi2r.r-13awgt0.r-18u37iz.r-1w6e6rj div.css-175oi2r.r-1mf7evn:nth-child(1) span.css-1qaijid.r-bcqeeo.r-qvutc0.r-poiln3.r-1b43r93.r-1cwl3u0.r-b88u0q span.css-1qaijid.r-bcqeeo.r-qvutc0.r-poiln3::text').extract_first()
            profile_followers = sel.css('div.css-175oi2r.r-ymttw5.r-ttdzmv.r-1ifxtd0 div.css-175oi2r.r-13awgt0.r-18u37iz.r-1w6e6rj div.css-175oi2r.r-1mf7evn:nth-child(2) span.css-1qaijid.r-bcqeeo.r-qvutc0.r-poiln3.r-1b43r93.r-1cwl3u0.r-b88u0q span.css-1qaijid.r-bcqeeo.r-qvutc0.r-poiln3::text').extract_first()
            
            twitter_data.append({
                "Name Twitter": profile_name.strip() if profile_name else None,
                "Profile Handle": profile_handle.strip() if profile_handle else None,
                "Profile Job Title": profile_job_title.strip() if profile_job_title else None,
                "Join Date": profile_joining_date.strip() if profile_joining_date else None,
                "Following Twitter": profile_following.strip() if profile_following else None,
                "Fowwers Twitter": profile_followers.strip() if profile_followers else None,
            })
        except Exception as e:
            print(f"Error extracting Twitter data: {str(e)}")

    return twitter_data

def search_social_media_profile(platform):
    input_file_path = file_entry.get()
    social_media_data =[]

    try:
        # Export the data to a new Excel file
        input_workbook = load_workbook(input_file_path, read_only=True)
        input_sheet = input_workbook.active
        driver = webdriver.Chrome()
        driver.get("https://www.google.com")
        time.sleep(5)
        lan = driver.find_element(By.XPATH, "/html/body/div[1]/div[4]/div/div/a[1]")
        if lan.text == "English":
            driver.get(lan.get_attribute("href"))
        # Create a list to store data in JSON format
        json_data = []
        i =0

        for row in input_sheet.iter_rows(min_row=2, values_only=True):
            if any(cell_value is not None and cell_value != "" for cell_value in row):
                i = i+1
                print(f"{i}")
                ceo_name = row[0] if row[0] else ""
                company_name = row[1] if row[1] else ""
                keywords = row[2] if row[2] else ""

                ceo_name = ""
                att = 0

                while not ceo_name:
                    query = f"{keywords} of {company_name} "
                    query = query.replace(" ", "%20")
                    search_url = f"https://www.google.com/search?q={query}"
                    driver.get(search_url)
                    time.sleep(5)
                    soup = BeautifulSoup(driver.page_source, "html.parser")
                    text = soup.get_text(strip=True)
                    check_ = ""
                    for _ in method_references:

                        try:
                            response = g4f.ChatCompletion.create(
                                model="gpt-3.5-turbo",
                                # provider=_,
                                messages=[
                                    {
                                        "role": "user",
                                        "content": f"I will give you a some text about a {keywords} of a company.This is the text :BEGIN: {text[:500]} :END. According the given text, your answer should be just the real {keywords} name,  your answer format answer will be :The {keywords} of Company A is B, no more explain. ",
                                    }
                                ],
                            )
                            check_ = "".join(messages for messages in response)
                            if (
                                "not" in check_
                                or not check_
                                or len(check_.split()) > 60
                                or "is" not in check_
                            ):
                                continue
                            else:
                                break
                        except:
                            continue

                    ceo_name = "".join(check_).replace("*", "")
                    print(ceo_name)
                    match = re.search(r'is\s+([^"]+)', ceo_name)
                    if match:
                        ceo_name = match.group(1)
                    ceo_name = ceo_name[:-1] if ceo_name.endswith(".") else ceo_name
                    nlp = spacy.load("./_internal/en_core_web_sm/en_core_web_sm-3.7.1")
                    # Process the text with SpaCy
                    doc = nlp(ceo_name)
                    # Extract person names
                    ceo_name = [ent.text for ent in doc.ents if ent.label_ == "PERSON"]
                    att += 1
                    if att == 3:
                        break
                
                ceo_name = "".join(ceo_name).strip()
                ceo_name = ceo_name.replace(".", "").replace("-", " ")
                if "not" in ceo_name:
                    ceo_name = f"{keywords} not found"
                    json_data.append(
                        {
                            "Name": ceo_name,
                            "Company Name": company_name,
                            "Keywords": keywords,
                            f"{platform} Profile": f"No {platform.lower()} links found.",
                            # "LinkedIn Data": [],
                        }
                    )
                else:
                    query = f"{platform.lower()} {ceo_name} {keywords} of {company_name}"
                    query = query.replace(" ", "%20")
                    search_url = f"https://www.google.com/search?q={query}"
                    driver.get(search_url)
                    time.sleep(5)
                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.END)
                    profile_links = []
                    name_profile = []
                    time.sleep(5)
                    element = driver.find_element(
                        By.CSS_SELECTOR, "div.eqAnXb > div#search > div > div"
                    )

                    # Get the HTML content of the element
                    element_html = element.get_attribute("outerHTML")

                    soup = BeautifulSoup(element_html, "html.parser")
                    # Extract href attributes

                    # Extract href attributes
                    # Extract the text or HTML of the element
                    for link, name in zip(
                        [
                            str(tag.get("href")).split("?")[0]
                            for tag in soup.find_all("a")
                        ],
                        [tag.text for tag in soup.find_all("h3")],
                    ):
                        # Or get the HTML of the element
                        if (
                            f"{platform.lower()}.com" in link
                            and "/status/" not in link
                            and "/post/" not in link
                            and "/posts/" not in link
                            and "post" not in link
                            and "posts" not in link
                            and "story" not in link
                            and "news" not in link
                            and "job" not in link
                            and "today" not in link
                            and "pulse" not in link
                            and "company" not in link
                            and "groups" not in link
                            and "text" not in link
                            and "translate" not in link
                            and "login" not in link
                            and "search" not in link
                            # and any(
                            #     word.lower() in name.lower()
                            #     for word in ceo_name.split()
                            # )
                        ):
                            print(link, name)
                            profile_links.append(link)
                            ceo_name = name.split("-")[0].strip()
                            name_profile.append(name)
                            break
                    if (
                        not profile_links
                        or "posts" in profile_links[-1]
                        or "/posts/" in profile_links[-1]
                    ):
                        json_data.append(
                            {
                                "Name": ceo_name,
                                "Company Name": company_name,
                                "Keywords": keywords,
                                f"{platform} Profile": f"No {platform.lower()} links found.",
                                "LinkedIn Data": [],
                                "Twitter Data": [],
                                # "Facebook Data": [],
                            }
                        )
                    else:
                        json_data.append(
                            {
                                "Name": ceo_name,
                                "Company Name": company_name,
                                "Keywords": keywords,
                                f"{platform} Profile": profile_links[-1],
                                "LinkedIn Data": [],
                                "Twitter Data": [],
                                # "Facebook Data": [],
                            }
                        )

                        if platform.lower() == "linkedin":
                            linkedin_data = extract_information_linkedin(driver, profile_links)
                            json_data[-1]["LinkedIn Data"] = linkedin_data

                        elif platform.lower() == "twitter":
                            twitter_data = extract_infomation_twitter(driver, profile_links)
                            json_data[-1]["Twitter Data"] = twitter_data

                        # elif platform.lower() == "facebook":
                        #     facebook_data = extract_information_facebook(driver, profile_links)
                        #     json_data[-1]["Facebook Data"] = facebook_data

        driver.quit()

        json_file_path = f"{platform.lower()}_data.json"
        with open(json_file_path, "w") as json_file:
            json.dump(json_data, json_file, indent=2)

        # Save JSON data to Excel
        excel_output_file_path = f"output_{platform.lower()}.xlsx"
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.append(
            [
                "Name",
                "Company Name",
                "Keywords",
                f"{platform.lower()} Profile",
                "Name LinkedIn",
                "Job Title",
                "Followers",
                "Location",
                "Name Twitter",
                "Profile Handle",
                "Profile Job Title",
                "Join Date",
                "Following Twitter",
                "Followers Twitter",
                # "Name Facebook",
                # "Followers Facebook",
                # "Following Facebook",
            ]
        )

        for entry in json_data:
            linkedin_data = entry.get("LinkedIn Data", [])
            twitter_data = entry.get("Twitter Data", [])
            # facebook_data = entry.get("Facebook Data", [])
            if linkedin_data:
                for linkedin_entry in linkedin_data:
                    output_sheet.append(
                        [
                            entry["Name"],
                            entry["Company Name"],
                            entry["Keywords"],
                            entry[f"{platform} Profile"],
                            linkedin_entry.get("Name LinkedIn", ""),
                            linkedin_entry.get("Job Title", ""),
                            linkedin_entry.get("Followers", ""),
                            linkedin_entry.get("Location", ""),
                            
                        ]
                    )

            elif twitter_data:
                for twitter_entry in twitter_data:
                    output_sheet.append(
                        [
                            entry["Name"],
                            entry["Company Name"],
                            entry["Keywords"],
                            entry[f"{platform} Profile"],
                            twitter_entry.get("Name Twitter", ""),
                            twitter_entry.get("Profile Handle", ""),
                            twitter_entry.get("Profile Job Title", ""),
                            twitter_entry.get("Join Date", ""),
                            twitter_entry.get("Following Twitter", ""),
                            twitter_entry.get("Fowwers Twitter", ""),
                        ]
                    )

            # elif facebook_data:
            #     for facebook_entry in facebook_data:
            #         output_sheet.append(
            #             [
            #                 entry["Name"],
            #                 entry["Company Name"],
            #                 entry["Keywords"],
            #                 entry[f"{platform} Profile"],
            #                 facebook_entry.get("Name And Company Facebook", ""),
            #                 facebook_entry.get("Followers Facebook", ""),
            #                 facebook_entry.get("Following Facebook", ""),
            #             ]
            #         )

            else:
                output_sheet.append([
                    entry["Name"],
                    entry["Company Name"],
                    entry["Keywords"],
                    entry[f"{platform} Profile"],
                ])

        output_workbook.save(excel_output_file_path)

        status_label.config(
            text=f"Output saved to: {excel_output_file_path} and {json_file_path}"
        )

    except Exception as e:
        error_message = f"Error: {str(e)}"
        status_label.config(text=error_message)

root = tk.Tk()
root.title("Find Social Media account")

# create the main frame
main_frame = tk.Frame(root, padx=10, pady=10)
main_frame.pack()

# create the file frame
file_frame = tk.LabelFrame(main_frame, text="Attached input file")
file_frame.pack(fill="x", padx=10, pady=10)

file_entry = tk.Entry(file_frame, width=40)
file_entry.pack(side="left", padx=10, pady=5)

file_button = tk.Button(file_frame, text="Browser", command=select_file)
file_button.pack(side="left", padx=10, pady=5)
# create the status label
status_label = tk.Label(main_frame, text="", font=("Arial", 12))
status_label.pack(pady=10)

for platform in SOCIAL_MEDIA_PLATFORMS:
    send_button = tk.Button(main_frame, text=f"Find {platform} Account", command=lambda p=platform: search_social_media_profile(p))
    send_button.pack(pady=5)

root.mainloop()